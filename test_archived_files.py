import io
import os
import zipfile
import csv
from openpyxl import load_workbook
from pypdf import PdfReader


def test_create_archive():
    if not os.path.exists("resources"):  # проверяем существует ли папка
        os.mkdir("resources")  # создаем папку если её нет
    archive_zip = "resources/archive.zip"
    with zipfile.ZipFile(archive_zip, 'w') as zf:  # создаем архив
        for file in os.listdir('tmp'):  # добавляем файлы в архив
            add_file = os.path.join('tmp', file)  # склеиваем путь к файлам которые добавляют в архив
            zf.write(add_file, os.path.basename(add_file))  # добавляем файл в архив


def test_file_exist_in_archive():
    expected_files = ["csv_sample.csv", "xlsx_sample.xlsx", "pdf_sample.pdf"]
    archive_zip = "resources/archive.zip"
    with zipfile.ZipFile(archive_zip, 'r') as zip_file:
        for file in expected_files:
            print(f"Файл {file} найден в архиве!")
            namelist_ = file in zip_file.namelist()
            assert namelist_


def test_content_csv():
    expected_data = [
        ['№', 'Column 1', 'Column 2', 'Column 3', 'Column 4', 'Column 5'],
        ['Row 1', '1', '1', '1', '1', '1'],
        ['Row 2', '2', '2', '2', '2', '2'],
        ['Row 3', '3', '3', '3', '3', '3'],
        ['Row 4', '4', '4', '4', '4', '4'],
        ['Row 5', '5', '5', '5', '5', '5']
    ]
    archive_zip = "resources/archive.zip"
    with zipfile.ZipFile(archive_zip, 'r') as zip_file:
        with zip_file.open('csv_sample.csv', 'r') as csv_file:
            text_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.reader(text_file)
            actual_data = list(reader)

            assert actual_data == expected_data


def test_content_xlsx():
    expected_data = [
        ['№', 'Column 1', 'Column 2', 'Column 3', 'Column 4', 'Column 5'],
        ['Row 1', 1.0, 1.0, 1.0, 1.0, 1.0],
        ['Row 2', 2.0, 2.0, 2.0, 2.0, 2.0],
        ['Row 3', 3.0, 3.0, 3.0, 3.0, 3.0],
        ['Row 4', 4.0, 4.0, 4.0, 4.0, 4.0],
        ['Row 5', 5.0, 5.0, 5.0, 5.0, 5.0]
    ]
    archive_zip = "resources/archive.zip"

    with zipfile.ZipFile(archive_zip, 'r') as zip_file:
        with zip_file.open('xlsx_sample.xlsx') as xlsx_file:
            workbook = load_workbook(filename=xlsx_file)
            sheet = workbook.active

            actual_data = []
            for row in sheet.iter_rows(values_only=True):
                actual_data.append(list(row))

            assert actual_data == expected_data


def test_content_pdf():
    expected_data = (
        '№ Column 1 Column 2 Column 3 Column 4 Column 5 \n'
        'Row 1 1 1 1 1 1 \n'
        'Row 2 2 2 2 2 2 \n'
        'Row 3 3 3 3 3 3 \n'
        'Row 4 4 4 4 4 4 \n'
        'Row 5 5 5 5 5 5 '
    )

    archive_zip = "resources/archive.zip"

    with zipfile.ZipFile(archive_zip, 'r') as zip_file:
        with zip_file.open('pdf_sample.pdf') as pdf_file:
            pdf_reader = PdfReader(io.BytesIO(pdf_file.read()))
            actual_data = ""
            for page in pdf_reader.pages:
                actual_data += page.extract_text()

            assert actual_data == expected_data
